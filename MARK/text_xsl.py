import pandas as pd
import time
import model
import numpy as np
from caculate import calculate_stats
from rouge import calculate_rouge_scores
from bleu import bleu
from openpyxl import load_workbook

def process_file(file_path, dropdown, socketio, filename, use_general_algorithm):
    # 读取Excel文件，假设文件从第二行开始有数据
    df = pd.read_excel(file_path, header=0)  # header=0 表示第一行为列名

    if '模型答案' not in df.columns:
        df['模型答案'] = None
    if '标准答案' not in df.columns:
        df['标准答案'] = None
    if '分数' not in df.columns:
        df['分数'] = None
    if '原因' not in df.columns:
        df['原因'] = None

    if dropdown == 'qwen2.5:3b':
        chat = model.chat_qwen
        moxing = 'qwen2.5_3b'
    else:
        chat = model.chat_deepseek
        moxing = 'deepseek-r1_1.5b'
    print(moxing)

    total_questions = len(df)
    stop_1 = 0
    for index, row in df.iterrows():
        if use_general_algorithm:
            a = pd.notna(row[1]) and pd.notna(row[2]) and pd.notna(row[3]) and pd.isna(row[4]) and pd.isna(row[5])
        else:
            a = pd.notna(row[1]) and pd.notna(row[2]) and pd.isna(row[4]) and pd.isna(row[5])
        
        if a: 
            question = row[1]
            answer = row[2]
            time.sleep(1)
            response = chat(question, answer)
            print("")
            fenshu, reason, stop = model.extract_score_and_reason(response)                
            if stop == 1:
                stop_1 = 1
                
            df.at[index, '分数'] = fenshu
            df.at[index, '原因'] = reason
                
            # 更新进度
            progress = (index + 1) / total_questions * 100
            socketio.emit('progress', {'filename': filename, 'progress': progress})
        else:
            if index == 0:
                error_message = f"文件格式错误: 您所选择的系统操作与您的文件格式不匹配"
            else:
                error_message = f"文件格式错误: 第 {index + 1} 行数据不完整。"
                
            socketio.emit('error', {'message': error_message})
            raise ValueError(error_message)  # 抛出异常以停止程序
    
    data = df.to_dict(orient='records')
    
    if use_general_algorithm:
        socketio.emit('status', {'message': '正在计算通用模型参数......'})
        bleu_1, bleu_2, bleu_3, bleu_4 = 0, 0, 0, 0
        if total_questions < 10:
            blue_liat = list(range(total_questions))
        else:
            blue_liat = np.random.choice(range(total_questions), size=10, replace=False).tolist()
        for i in blue_liat:
            bleu_score = bleu(data[i]['标准答案'], data[i]['模型答案'])
            bleu_1 = bleu_1 + bleu_score[0]
            bleu_2 = bleu_2 + bleu_score[1]
            bleu_3 = bleu_3 + bleu_score[2]
            bleu_4 = bleu_4 + bleu_score[3]
            bleu_score.clear()
        
        rouge = calculate_rouge_scores(df)
    
    numbers = [item['分数'] for item in data]
    average, medium, standard_deviation = calculate_stats(numbers) 
    numbers.clear()

    result_file_path = file_path.replace('.xlsx', '_'+moxing+'_t_r.xlsx') 
    
    with pd.ExcelWriter(result_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='数据', index=False)
        
        # 获取工作簿和工作表
        workbook = writer.book
        stats_sheet = workbook.create_sheet(title='统计')

        # 写入统计信息到指定单元格
        stats_sheet['A1'] = '总题数'
        stats_sheet['A2'] = total_questions
        stats_sheet['B1'] = '平均分'
        stats_sheet['B2'] = average
        stats_sheet['C1'] = '中位数'
        stats_sheet['C2'] = medium
        stats_sheet['D1'] = '标准差'
        stats_sheet['D2'] = standard_deviation
        stats_sheet['A4'] = '不计算通用模型参数'
        
        if use_general_algorithm:
            stats_sheet['A4'] = 'bleu-1'
            stats_sheet['A5'] = bleu_1 / len(blue_liat)
            stats_sheet['B4'] = 'bleu-2'
            stats_sheet['B5'] = bleu_2 / len(blue_liat)
            stats_sheet['C4'] = 'bleu-3'
            stats_sheet['C5'] = bleu_3 / len(blue_liat)
            stats_sheet['D4'] = 'bleu-4'
            stats_sheet['D5'] = bleu_4 / len(blue_liat)
            rouge_sheet = workbook.create_sheet(title='rouge')
            ROUGE = pd.DataFrame(rouge).T
            ROUGE.to_excel(writer, sheet_name='rouge', index=True)

    if stop_1 == 1:  # 如果有停止标志，则发送停止信号
        socketio.emit('status', {'message': '回答中有错误，请检查结果！(-9999)'})
    else:
        socketio.emit('status', {'message': '回答成功!'})